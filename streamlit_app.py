"""
Streamlit Web Application - Champion Cleaners Google Ads Bot
Exact replica of Flask localhost app on Streamlit Cloud
"""

import streamlit as st
import pandas as pd
import numpy as np
import json
from pathlib import Path
import sys
from datetime import datetime

# Add paths
sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, str(Path(__file__).parent / 'keyword_engine_v2'))

try:
    from main_windows import ChampionCleanersBot
    from keyword_main import KeywordIntelligenceEngine
except ImportError as e:
    st.error(f"Error loading modules: {e}")
    sys.exit(1)

# Helper function to convert Google Ads keyword format to required format
def convert_google_ads_keywords(df):
    """Convert Google Ads keyword report format to standard format."""
    try:
        # Mapping of Google Ads columns to required columns
        column_mapping = {
            'Keyword': 'keyword',
            'Match type': 'match_type',
            'Clicks': 'clicks',
            'Impr.': 'impressions',
            'Cost': 'cost',
            'Conversions': 'conversions'
        }
        
        # Rename columns
        df_converted = df.copy()
        for old_col, new_col in column_mapping.items():
            if old_col in df_converted.columns:
                df_converted.rename(columns={old_col: new_col}, inplace=True)
        
        # Add required columns if missing
        if 'campaign_name' not in df_converted.columns:
            df_converted['campaign_name'] = 'Keywords Report'
        if 'ad_group_name' not in df_converted.columns:
            df_converted['ad_group_name'] = 'Keyword Group'
        
        # Keep only required columns
        required_cols = ['campaign_name', 'ad_group_name', 'keyword', 'match_type', 'clicks', 'impressions', 'cost', 'conversions']
        df_converted = df_converted[[col for col in required_cols if col in df_converted.columns]]
        
        # Clean numeric columns
        for col in ['clicks', 'impressions', 'cost', 'conversions']:
            if col in df_converted.columns:
                # Remove commas and convert to numeric
                df_converted[col] = df_converted[col].astype(str).str.replace(',', '')
                df_converted[col] = pd.to_numeric(df_converted[col], errors='coerce').fillna(0)
        
        return df_converted
    except Exception as e:
        st.error(f"Error converting Google Ads format: {str(e)}")
        return None

# Helper function to convert Google Ads campaign format to required format
def convert_google_ads_campaigns(df):
    """Convert Google Ads campaign report format to standard format."""
    try:
        # Mapping of Google Ads campaign columns to required columns
        column_mapping = {
            'Campaign': 'campaign_name',
            'Campaign type': 'campaign_type',
            'Impr.': 'impressions',
            'Interactions': 'clicks',
            'Cost': 'cost',
            'Conversions': 'conversions'
        }
        
        # Rename columns
        df_converted = df.copy()
        for old_col, new_col in column_mapping.items():
            if old_col in df_converted.columns:
                df_converted.rename(columns={old_col: new_col}, inplace=True)
        
        # Add date column if missing (use report date or default)
        if 'date' not in df_converted.columns:
            df_converted['date'] = pd.Timestamp.now().strftime('%Y-%m-%d')
        
        # Keep only required columns
        required_cols = ['date', 'campaign_name', 'campaign_type', 'clicks', 'impressions', 'cost', 'conversions']
        df_converted = df_converted[[col for col in required_cols if col in df_converted.columns]]
        
        # Clean numeric columns - remove commas and convert
        for col in ['clicks', 'impressions', 'cost', 'conversions']:
            if col in df_converted.columns:
                df_converted[col] = df_converted[col].astype(str).str.replace(',', '')
                df_converted[col] = pd.to_numeric(df_converted[col], errors='coerce').fillna(0)
        
        # Remove totals rows (any row where campaign_name contains "Total")
        if 'campaign_name' in df_converted.columns:
            df_converted = df_converted[~df_converted['campaign_name'].astype(str).str.contains('Total', case=False, na=False)]
        
        return df_converted
    except Exception as e:
        st.error(f"Error converting Google Ads campaign format: {str(e)}")
        return None

# Configure page
st.set_page_config(
    page_title="Champion Cleaners - Google Ads Bot",
    page_icon="🤖",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS matching localhost
st.markdown("""
<style>
    .main {
        padding: 2rem;
    }
    [data-testid="stMetricValue"] {
        font-size: 2em;
        font-weight: bold;
        color: #667eea;
    }
    .stTabs [data-baseweb="tab-list"] button [data-testid="stMarkdownContainer"] p {
        font-size: 1em;
        font-weight: 600;
    }
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        padding: 20px;
        border-radius: 10px;
        margin: 10px 0;
    }
</style>
""", unsafe_allow_html=True)

# Initialize session state
if 'analysis_type' not in st.session_state:
    st.session_state.analysis_type = 'campaign'
if 'campaign_results' not in st.session_state:
    st.session_state.campaign_results = None
if 'keyword_results' not in st.session_state:
    st.session_state.keyword_results = None

# Header
st.markdown("""
    <div style="text-align: center; margin-bottom: 40px;">
        <h1 style="color: #667eea;">🤖 Champion Cleaners - Google Ads Bot</h1>
        <p style="font-size: 1.1em; color: #666;">AI-Powered Google Ads Analysis & Optimization</p>
    </div>
""", unsafe_allow_html=True)

# Main tabs for analysis type
col1, col2 = st.columns(2)
with col1:
    if st.button("📊 Campaign Analysis", key="campaign_btn", use_container_width=True):
        st.session_state.analysis_type = 'campaign'
        st.rerun()

with col2:
    if st.button("🔍 Keyword Engine", key="keyword_btn", use_container_width=True):
        st.session_state.analysis_type = 'keyword'
        st.rerun()

st.markdown("---")

# ==================== CAMPAIGN ANALYSIS ====================
if st.session_state.analysis_type == 'campaign':
    st.subheader("📊 Campaign Analysis")
    
    # Tabs for input
    tab1, tab2 = st.tabs(["Sample Data", "Upload CSV"])
    
    with tab1:
        st.write("Analyze using sample campaign data")
        if st.button("▶️ Run Sample Analysis", key="campaign_sample_btn", use_container_width=True):
            with st.spinner("Analyzing sample data..."):
                try:
                    sample_path = Path(__file__).parent / 'sample_data.csv'
                    if sample_path.exists():
                        bot = ChampionCleanersBot(str(sample_path), use_emojis=False)
                        results = bot.run_analysis(verbose=False)
                        st.session_state.campaign_results = results
                        st.success("✅ Analysis complete!")
                    else:
                        st.error("Sample data file not found")
                except Exception as e:
                    st.error(f"Error: {str(e)}")
    
    with tab2:
        st.write("Upload your CSV file for analysis")
        st.info("📌 Supports both standard CSV format and Google Ads campaign reports")
        uploaded_file = st.file_uploader("Choose CSV file", type=['csv'], key="campaign_upload")
        if uploaded_file is not None:
            if st.button("▶️ Run Analysis", key="campaign_upload_btn", use_container_width=True):
                with st.spinner("Analyzing data..."):
                    try:
                        upload_path = Path(__file__).parent / 'uploads' / uploaded_file.name
                        upload_path.parent.mkdir(exist_ok=True)
                        
                        # Save file with proper encoding
                        content = uploaded_file.read()
                        with open(upload_path, 'wb') as f:
                            f.write(content)
                        
                        # Try to load the file
                        conversion_failed = False
                        df_loaded = None
                        
                        # Strategy 1: Try standard format
                        try:
                            df_loaded = pd.read_csv(upload_path)
                        except Exception as e1:
                            st.write(f"Standard format failed: {str(e1)}")
                            
                            # Strategy 2: Try skipping 2 rows (Google Ads format)
                            try:
                                df_loaded = pd.read_csv(upload_path, skiprows=2)
                            except Exception as e2:
                                st.write(f"Skiprows 2 failed: {str(e2)}")
                                
                                # Strategy 3: Try skipping 3 rows
                                try:
                                    df_loaded = pd.read_csv(upload_path, skiprows=3)
                                except Exception as e3:
                                    st.error(f"All parsing strategies failed. Last error: {str(e3)}")
                                    conversion_failed = True
                        
                        if df_loaded is not None and not conversion_failed:
                            # Check if it's Google Ads campaign format
                            if 'Campaign' in df_loaded.columns or 'Impr.' in df_loaded.columns:
                                st.write("✅ Detected Google Ads campaign format - converting...")
                                df_converted = convert_google_ads_campaigns(df_loaded)
                                if df_converted is None:
                                    conversion_failed = True
                                else:
                                    df_loaded = df_converted
                            
                            if not conversion_failed:
                                # Save converted file
                                df_loaded.to_csv(upload_path, index=False)
                                st.write(f"✅ File loaded: {len(df_loaded)} campaigns")
                        
                        if not conversion_failed and df_loaded is not None:
                            bot = ChampionCleanersBot(str(upload_path), use_emojis=False)
                            results = bot.run_analysis(verbose=False)
                            
                            if results:
                                st.session_state.campaign_results = results
                                st.success("✅ Analysis complete!")
                            else:
                                st.error("Analysis failed - check file format and data")
                    except Exception as e:
                        st.error(f"Error: {str(e)}")
                        import traceback
                        st.write(traceback.format_exc())
    
    # Display results
    if st.session_state.campaign_results:
        results = st.session_state.campaign_results
        st.markdown("---")
        st.subheader("📈 Analysis Results")
        
        # Results tabs
        res_tab1, res_tab2, res_tab3, res_tab4, res_tab5, res_tab6 = st.tabs([
            "Campaign Metrics",
            "Platform Analysis",
            "Device Analysis",
            "Issues Detected",
            "Recommendations",
            "Budget Allocation"
        ])
        
        with res_tab1:
            st.write("### Campaign Metrics")
            if 'campaign_metrics' in results:
                metrics = results['campaign_metrics']
                
                # Show each campaign's metrics
                if isinstance(metrics, dict):
                    for campaign_name, campaign_data in metrics.items():
                        st.subheader(f"{campaign_name}")
                        
                        if isinstance(campaign_data, dict):
                            col1, col2, col3, col4 = st.columns(4)
                            
                            items = list(campaign_data.items())
                            for idx, (key, value) in enumerate(items[:4]):
                                col = [col1, col2, col3, col4][idx % 4]
                                with col:
                                    if isinstance(value, (int, float)):
                                        st.metric(key.replace('_', ' ').title(), f"{value:,.2f}" if isinstance(value, float) else f"{value:,}")
                                    else:
                                        st.metric(key.replace('_', ' ').title(), str(value))
                            
                            # Show remaining metrics if any
                            if len(items) > 4:
                                with st.expander("📊 More Metrics"):
                                    metrics_df = pd.DataFrame([campaign_data])
                                    st.dataframe(metrics_df, use_container_width=True)
        
        with res_tab2:
            st.write("### Platform Analysis")
            if 'platform_analysis' in results:
                platform = results['platform_analysis']
                
                if platform and isinstance(platform, dict):
                    for platform_name, platform_data in platform.items():
                        st.subheader(f"{platform_name}")
                        
                        if isinstance(platform_data, dict):
                            col1, col2, col3 = st.columns(3)
                            
                            items = list(platform_data.items())
                            for idx, (key, value) in enumerate(items[:3]):
                                col = [col1, col2, col3][idx % 3]
                                with col:
                                    if isinstance(value, (int, float)):
                                        st.metric(key.replace('_', ' ').title(), f"{value:,.2f}" if isinstance(value, float) else f"{value:,}")
                                    else:
                                        st.metric(key.replace('_', ' ').title(), str(value))
                            
                            # Show remaining data if any
                            if len(items) > 3:
                                with st.expander("📊 More Platform Data"):
                                    platform_df = pd.DataFrame([platform_data])
                                    st.dataframe(platform_df, use_container_width=True)
                else:
                    st.info("No platform analysis data available")
        
        with res_tab3:
            st.write("### Device Analysis")
            if 'device_analysis' in results:
                device = results['device_analysis']
                
                if device and isinstance(device, dict):
                    for device_name, device_data in device.items():
                        st.subheader(f"{device_name}")
                        
                        if isinstance(device_data, dict):
                            col1, col2, col3 = st.columns(3)
                            
                            items = list(device_data.items())
                            for idx, (key, value) in enumerate(items[:3]):
                                col = [col1, col2, col3][idx % 3]
                                with col:
                                    if isinstance(value, (int, float)):
                                        st.metric(key.replace('_', ' ').title(), f"{value:,.2f}" if isinstance(value, float) else f"{value:,}")
                                    else:
                                        st.metric(key.replace('_', ' ').title(), str(value))
                            
                            # Show remaining data if any
                            if len(items) > 3:
                                with st.expander("📊 More Device Data"):
                                    device_df = pd.DataFrame([device_data])
                                    st.dataframe(device_df, use_container_width=True)
                else:
                    st.info("No device analysis data available")
        
        with res_tab4:
            st.write("### Issues Detected")
            if 'detected_issues' in results:
                issues = results['detected_issues']
                if isinstance(issues, list) and len(issues) > 0:
                    for i, issue in enumerate(issues, 1):
                        severity = issue.get('severity', 'Medium')
                        campaign = issue.get('campaign', 'Unknown')
                        description = issue.get('description', 'No description')
                        
                        icon = "🔴" if severity == "High" else "🟡"
                        st.warning(f"**{icon} {i}. [{severity}] {campaign}**\n\n{description}")
                else:
                    st.success("No issues detected ✅")
        
        with res_tab5:
            st.write("### Recommendations")
            if 'recommendations' in results:
                recommendations = results['recommendations']
                if isinstance(recommendations, list) and len(recommendations) > 0:
                    for i, rec in enumerate(recommendations, 1):
                        campaign_name = rec.get('campaign_name', 'Unknown')
                        issue = rec.get('issue_detected', 'No issue')
                        recommendation = rec.get('recommendation', 'No recommendation')
                        confidence = rec.get('confidence_level', 'Medium')
                        
                        icon = "🔴" if confidence == "High" else "🟡"
                        st.info(f"**{icon} {i}. {campaign_name}**\n\n**Issue:** {issue}\n\n**Recommendation:** {recommendation}\n\n**Confidence:** {confidence}")
                else:
                    st.info("No recommendations available")
        
        with res_tab6:
            st.write("### Budget Allocation")
            if 'budget_allocation' in results:
                budget = results['budget_allocation']
                
                if isinstance(budget, dict):
                    # Show total budget
                    if 'total_monthly_budget' in budget:
                        st.metric("Total Monthly Budget", f"AED {budget['total_monthly_budget']:,.2f}")
                    
                    # Show allocations
                    if 'allocations' in budget and isinstance(budget['allocations'], dict):
                        st.subheader("Campaign Budget Allocation")
                        
                        alloc_data = []
                        for campaign, alloc in budget['allocations'].items():
                            current_pct = alloc.get('current_percentage', 0)
                            rec_pct = alloc.get('recommended_percentage', current_pct)
                            adjust = alloc.get('budget_adjustment', 0)
                            
                            alloc_data.append({
                                'Campaign': campaign,
                                'Current %': f"{current_pct:.1f}%",
                                'Recommended %': f"{rec_pct:.1f}%",
                                'Adjustment': f"{adjust:+.1f}%" if adjust != 0 else "No change"
                            })
                        
                        if alloc_data:
                            alloc_df = pd.DataFrame(alloc_data)
                            st.dataframe(alloc_df, use_container_width=True, hide_index=True)
        
        # Export buttons
        st.markdown("---")
        col1, col2 = st.columns(2)
        with col1:
            json_str = json.dumps(results, indent=2, default=str)
            st.download_button(
                label="📥 Download JSON",
                data=json_str,
                file_name=f"analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json"
            )
        
        with col2:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                
                # Create Excel workbook
                wb = Workbook()
                
                # Remove default sheet
                wb.remove(wb.active)
                
                # Define styles
                header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=12)
                
                # Summary sheet
                ws_summary = wb.create_sheet('Summary')
                ws_summary['A1'] = 'Campaign Analysis Summary'
                ws_summary['A1'].font = Font(bold=True, size=14)
                
                row = 3
                if 'data_summary' in results:
                    summary = results['data_summary']
                    for key, value in summary.items():
                        ws_summary[f'A{row}'] = key.replace('_', ' ').title()
                        ws_summary[f'B{row}'] = value
                        row += 1
                
                # Recommendations sheet
                if 'recommendations' in results:
                    ws_rec = wb.create_sheet('Recommendations')
                    
                    headers = ['Campaign', 'Issue', 'Recommendation', 'Confidence']
                    for col_num, header in enumerate(headers, 1):
                        cell = ws_rec.cell(row=1, column=col_num)
                        cell.value = header
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    for row_num, rec in enumerate(results['recommendations'], 2):
                        ws_rec.cell(row=row_num, column=1).value = rec.get('campaign_name', '')
                        ws_rec.cell(row=row_num, column=2).value = rec.get('issue_detected', '')
                        ws_rec.cell(row=row_num, column=3).value = rec.get('recommendation', '')
                        ws_rec.cell(row=row_num, column=4).value = rec.get('confidence_level', '')
                
                # Issues sheet
                if 'detected_issues' in results:
                    ws_issues = wb.create_sheet('Issues')
                    
                    headers = ['Campaign', 'Severity', 'Description']
                    for col_num, header in enumerate(headers, 1):
                        cell = ws_issues.cell(row=1, column=col_num)
                        cell.value = header
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    for row_num, issue in enumerate(results['detected_issues'], 2):
                        ws_issues.cell(row=row_num, column=1).value = issue.get('campaign', '')
                        ws_issues.cell(row=row_num, column=2).value = issue.get('severity', '')
                        ws_issues.cell(row=row_num, column=3).value = issue.get('description', '')
                
                # Save to bytes
                from io import BytesIO
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                st.download_button(
                    label="📊 Download Excel",
                    data=output.getvalue(),
                    file_name=f"analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except ImportError:
                st.error("Excel export requires openpyxl package")

# ==================== KEYWORD ENGINE ====================
else:
    st.subheader("🔍 Keyword Intelligence Engine")
    
    # Tabs for input
    tab1, tab2 = st.tabs(["Sample Keywords", "Upload Keywords"])
    
    with tab1:
        st.write("Analyze using sample keywords data")
        if st.button("▶️ Run Keyword Analysis", key="keyword_sample_btn", use_container_width=True):
            with st.spinner("Analyzing sample keywords..."):
                try:
                    sample_path = Path(__file__).parent / 'sample_keywords.csv'
                    if sample_path.exists():
                        engine = KeywordIntelligenceEngine()
                        if engine.load_keywords(str(sample_path)):
                            if engine.run_full_analysis():
                                results = engine.get_results_summary()
                                st.session_state.keyword_results = results
                                st.success("✅ Keyword analysis complete!")
                            else:
                                st.error("Analysis failed - check file format")
                        else:
                            st.error("Failed to load sample keywords - check file format")
                    else:
                        st.error("Sample keywords file not found")
                except Exception as e:
                    st.error(f"Error: {str(e)}")
                    import traceback
                    st.write(traceback.format_exc())
    
    with tab2:
        st.write("Upload your keywords CSV file")
        st.info("📌 Supports both standard CSV format and Google Ads keyword reports")
        uploaded_file = st.file_uploader("Choose CSV file", type=['csv'], key="keyword_upload")
        if uploaded_file is not None:
            if st.button("▶️ Run Keyword Analysis", key="keyword_upload_btn", use_container_width=True):
                with st.spinner("Analyzing keywords..."):
                    try:
                        # Save uploaded file with proper handling
                        upload_path = Path(__file__).parent / 'uploads' / uploaded_file.name
                        upload_path.parent.mkdir(exist_ok=True)
                        
                        # Read and save the file with UTF-8 encoding
                        content = uploaded_file.read()
                        with open(upload_path, 'wb') as f:
                            f.write(content)
                        
                        # Try to load the file
                        conversion_failed = False
                        df_loaded = None
                        
                        # Strategy 1: Try standard format
                        try:
                            df_loaded = pd.read_csv(upload_path)
                        except Exception as e1:
                            st.write(f"Standard format failed: {str(e1)}")
                            
                            # Strategy 2: Try skipping 2 rows (Google Ads format)
                            try:
                                df_loaded = pd.read_csv(upload_path, skiprows=2)
                            except Exception as e2:
                                st.write(f"Skiprows 2 failed: {str(e2)}")
                                
                                # Strategy 3: Try skipping 3 rows
                                try:
                                    df_loaded = pd.read_csv(upload_path, skiprows=3)
                                except Exception as e3:
                                    st.error(f"All parsing strategies failed. Last error: {str(e3)}")
                                    conversion_failed = True
                        
                        if df_loaded is not None and not conversion_failed:
                            # Check if it's Google Ads format
                            if 'Keyword' in df_loaded.columns or 'Impr.' in df_loaded.columns:
                                st.write("✅ Detected Google Ads format - converting...")
                                df_converted = convert_google_ads_keywords(df_loaded)
                                if df_converted is None:
                                    conversion_failed = True
                                else:
                                    df_loaded = df_converted
                            
                            if not conversion_failed:
                                # Save converted file
                                df_loaded.to_csv(upload_path, index=False)
                                st.write(f"✅ File loaded: {len(df_loaded)} keywords")
                        
                        if not conversion_failed and df_loaded is not None:
                            engine = KeywordIntelligenceEngine()
                            if engine.load_keywords(str(upload_path)):
                                if engine.run_full_analysis():
                                    results = engine.get_results_summary()
                                    st.session_state.keyword_results = results
                                    st.success("✅ Keyword analysis complete!")
                                else:
                                    st.error("Analysis failed during processing")
                            else:
                                st.error("Failed to load keyword file after conversion")
                    except Exception as e:
                        st.error(f"Error: {str(e)}")
                        import traceback
                        st.write(traceback.format_exc())
    
    # Display results
    if st.session_state.keyword_results:
        results = st.session_state.keyword_results
        st.markdown("---")
        st.subheader("📊 Keyword Analysis Results")
        
        # Summary metrics
        summary = results.get('summary', {})
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Total Keywords", summary.get('total_keywords', 0))
        with col2:
            st.metric("Keywords with Issues", summary.get('keywords_with_issues', 0))
        with col3:
            st.metric("Lost Opportunities", summary.get('lost_search_opportunities', 0))
        with col4:
            st.metric("New Keywords to Add", summary.get('new_keywords_suggested', 0))
        
        st.markdown("---")
        
        # Results tabs
        kw_tab1, kw_tab2, kw_tab3, kw_tab4, kw_tab5, kw_tab6 = st.tabs([
            "Summary",
            "Keyword Audit",
            "Lost Searches",
            "Match Types",
            "Alignment",
            "Recommendations"
        ])
        
        # Summary Tab
        with kw_tab1:
            st.write("### Summary Overview")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Keywords", summary.get('total_keywords', 0))
            with col2:
                st.metric("Match Type Recommendations", summary.get('match_type_conversions_recommended', 0))
            with col3:
                st.metric("Total Recommendations", summary.get('total_recommendations', 0))
        
        # Keyword Audit Tab
        with kw_tab2:
            st.write("### Keyword Health Audit Issues")
            audit = results.get('audit', [])
            if audit and isinstance(audit, list):
                audit_data = []
                for issue in audit[:20]:
                    audit_data.append({
                        'Keyword': issue.get('keyword', 'N/A'),
                        'Campaign': issue.get('campaign', 'N/A'),
                        'Issue Type': issue.get('issue_type', 'N/A'),
                        'Severity': issue.get('severity', 'N/A'),
                        'Value': str(issue.get('value', 'N/A'))
                    })
                if audit_data:
                    audit_df = pd.DataFrame(audit_data)
                    st.dataframe(audit_df, use_container_width=True, hide_index=True)
            else:
                st.info("No audit issues found")
        
        # Lost Searches Tab
        with kw_tab3:
            st.write("### Lost Search Opportunities")
            lost = results.get('lost_searches', [])
            if lost and isinstance(lost, list):
                lost_data = []
                for item in lost[:20]:
                    lost_data.append({
                        'Keyword': item.get('keyword', 'N/A'),
                        'Campaign': item.get('campaign', 'N/A'),
                        'Match Type': item.get('match_type', 'N/A'),
                        'Loss Type': item.get('loss_type', 'N/A'),
                        'Lost Searches': item.get('potential_searches_lost', 0),
                        'Recommendation': item.get('recommendation', 'N/A')[:50]
                    })
                if lost_data:
                    lost_df = pd.DataFrame(lost_data)
                    st.dataframe(lost_df, use_container_width=True, hide_index=True)
            else:
                st.info("No lost search opportunities found")
        
        # Match Types Tab
        with kw_tab4:
            st.write("### Match Type Optimization Recommendations")
            match_recs = results.get('match_recommendations', [])
            if match_recs and isinstance(match_recs, list):
                match_data = []
                for item in match_recs[:20]:
                    match_data.append({
                        'Keyword': item.get('keyword', 'N/A'),
                        'Campaign': item.get('campaign', 'N/A'),
                        'Current': item.get('current_match_type', 'N/A'),
                        'Recommended': item.get('recommended_match_type', 'N/A'),
                        'Reason': item.get('reason', 'N/A')[:40],
                        'Expected Impact': item.get('expected_impact', 'N/A'),
                        'Confidence': item.get('confidence', 'N/A')
                    })
                if match_data:
                    match_df = pd.DataFrame(match_data)
                    st.dataframe(match_df, use_container_width=True, hide_index=True)
            else:
                st.info("No match type recommendations")
        
        # Alignment Tab
        with kw_tab5:
            st.write("### Website-Keyword Alignment Analysis")
            alignment = results.get('alignment_analysis', [])
            if alignment and isinstance(alignment, list):
                alignment_data = []
                for item in alignment[:20]:
                    alignment_data.append({
                        'Keyword': item.get('keyword', 'N/A'),
                        'Status': item.get('status', 'N/A'),
                        'Relevance Score': f"{item.get('relevance_score', 0):.0f}%",
                        'Suggestions': item.get('suggestions', 'None')[:50]
                    })
                if alignment_data:
                    alignment_df = pd.DataFrame(alignment_data)
                    st.dataframe(alignment_df, use_container_width=True, hide_index=True)
            else:
                st.info("No alignment analysis available")
        
        # Recommendations Tab
        with kw_tab6:
            st.write("### Actionable Recommendations")
            recs = results.get('recommendations', [])
            if recs and isinstance(recs, list):
                for i, rec in enumerate(recs[:15], 1):
                    keyword = rec.get('keyword', 'N/A')
                    priority = rec.get('priority', 'Medium')
                    action = rec.get('action', 'No action')
                    problem = rec.get('problem', 'No problem stated')
                    
                    icon = "🔴" if priority == "High" else "🟡"
                    st.info(f"**{icon} {i}. {keyword}**\n\n**Priority:** {priority}\n\n**Problem:** {problem}\n\n**Action:** {action}")
            else:
                st.info("No recommendations available")
        
        # Export buttons
        st.markdown("---")
        col1, col2 = st.columns(2)
        with col1:
            json_str = json.dumps(results, indent=2, default=str)
            st.download_button(
                label="📥 Download JSON",
                data=json_str,
                file_name=f"keyword_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json"
            )
        
        with col2:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill, Font, Alignment
                
                wb = Workbook()
                wb.remove(wb.active)
                
                # Define styles
                header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF', size=12)
                title_font = Font(bold=True, size=14)
                
                # Summary sheet
                ws = wb.create_sheet('Summary', 0)
                ws['A1'] = 'Keyword Analysis Report'
                ws['A1'].font = title_font
                ws.merge_cells('A1:B1')
                
                row = 3
                summary_metrics = [
                    ('Total Keywords', summary.get('total_keywords', 0)),
                    ('Keywords with Issues', summary.get('keywords_with_issues', 0)),
                    ('Lost Opportunities', summary.get('lost_search_opportunities', 0)),
                    ('Match Type Conversions', summary.get('match_type_conversions_recommended', 0)),
                    ('New Keywords Suggested', summary.get('new_keywords_suggested', 0)),
                    ('Total Recommendations', summary.get('total_recommendations', 0))
                ]
                
                for metric, value in summary_metrics:
                    ws[f'A{row}'] = metric
                    ws[f'A{row}'].font = Font(bold=True)
                    ws[f'B{row}'] = value
                    ws[f'B{row}'].alignment = Alignment(horizontal='center')
                    row += 1
                
                ws.column_dimensions['A'].width = 25
                ws.column_dimensions['B'].width = 15
                
                # Keyword Audit sheet
                if results.get('audit'):
                    ws = wb.create_sheet('Keyword Audit')
                    headers = ['Keyword', 'Campaign', 'Issue Type', 'Severity', 'Value']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col)
                        cell.value = header
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    for row_idx, item in enumerate(results.get('audit', [])[:100], 2):
                        ws.cell(row=row_idx, column=1).value = item.get('keyword', '')
                        ws.cell(row=row_idx, column=2).value = item.get('campaign', '')
                        ws.cell(row=row_idx, column=3).value = item.get('issue_type', '')
                        ws.cell(row=row_idx, column=4).value = item.get('severity', '')
                        ws.cell(row=row_idx, column=5).value = item.get('value', '')
                
                # Lost Searches sheet
                if results.get('lost_searches'):
                    ws = wb.create_sheet('Lost Searches')
                    headers = ['Keyword', 'Campaign', 'Match Type', 'Loss Type', 'Lost Searches']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col)
                        cell.value = header
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    for row_idx, item in enumerate(results.get('lost_searches', [])[:100], 2):
                        ws.cell(row=row_idx, column=1).value = item.get('keyword', '')
                        ws.cell(row=row_idx, column=2).value = item.get('campaign', '')
                        ws.cell(row=row_idx, column=3).value = item.get('match_type', '')
                        ws.cell(row=row_idx, column=4).value = item.get('loss_type', '')
                        ws.cell(row=row_idx, column=5).value = item.get('potential_searches_lost', 0)
                
                # Match Recommendations sheet
                if results.get('match_recommendations'):
                    ws = wb.create_sheet('Match Types')
                    headers = ['Keyword', 'Campaign', 'Current Type', 'Recommended Type', 'Reason', 'Expected Impact', 'Confidence']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col)
                        cell.value = header
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    for row_idx, item in enumerate(results.get('match_recommendations', [])[:100], 2):
                        ws.cell(row=row_idx, column=1).value = item.get('keyword', '')
                        ws.cell(row=row_idx, column=2).value = item.get('campaign', '')
                        ws.cell(row=row_idx, column=3).value = item.get('current_match_type', '')
                        ws.cell(row=row_idx, column=4).value = item.get('recommended_match_type', '')
                        ws.cell(row=row_idx, column=5).value = item.get('reason', '')
                        ws.cell(row=row_idx, column=6).value = item.get('expected_impact', '')
                        ws.cell(row=row_idx, column=7).value = item.get('confidence', '')
                
                # Recommendations sheet
                if results.get('recommendations'):
                    ws = wb.create_sheet('Recommendations')
                    headers = ['Keyword', 'Priority', 'Problem', 'Action']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col)
                        cell.value = header
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    for row_idx, item in enumerate(results.get('recommendations', [])[:100], 2):
                        ws.cell(row=row_idx, column=1).value = item.get('keyword', '')
                        ws.cell(row=row_idx, column=2).value = item.get('priority', '')
                        ws.cell(row=row_idx, column=3).value = item.get('problem', '')
                        ws.cell(row=row_idx, column=4).value = item.get('action', '')
                
                # Save to bytes
                from io import BytesIO
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                st.download_button(
                    label="📊 Download Excel",
                    data=output.getvalue(),
                    file_name=f"keyword_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except ImportError:
                st.error("Excel export requires openpyxl package")

# Footer
st.markdown("---")
col1, col2, col3 = st.columns(3)
with col1:
    st.markdown("**📚 [Documentation](https://github.com/adeelciit786-hue/CCAD)**")
with col2:
    st.markdown("**🐛 [Report Issues](https://github.com/adeelciit786-hue/CCAD/issues)**")
with col3:
    st.markdown("**⭐ [GitHub](https://github.com/adeelciit786-hue/CCAD)**")

st.caption(f"Champion Cleaners Google Ads Bot v2.0.0 | {datetime.now().strftime('%B %d, %Y')}")
