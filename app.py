"""
Flask Web Application for Champion Cleaners Bot
Provides a web interface to run analyses and view recommendations
"""

from flask import Flask, render_template, request, jsonify, send_file
import os
import json
from pathlib import Path
import sys
from io import BytesIO

# Ensure we use the venv packages
import site
site_packages = Path(__file__).parent / '.venv' / 'Lib' / 'site-packages'
if str(site_packages) not in sys.path:
    sys.path.insert(0, str(site_packages))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Add src directory to path
sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, str(Path(__file__).parent / 'keyword_engine_v2'))

from main_windows import ChampionCleanersBot

try:
    from keyword_main import KeywordIntelligenceEngine
    from website_relevance_checker import WebsiteRelevanceChecker
except ImportError as e:
    # Fallback if keyword modules not available
    KeywordIntelligenceEngine = None
    WebsiteRelevanceChecker = None
    print(f"Warning: Keyword modules not fully available: {e}")

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
UPLOAD_FOLDER = Path(__file__).parent / 'uploads'
UPLOAD_FOLDER.mkdir(exist_ok=True)

@app.route('/')
def index():
    """Render home page"""
    return render_template('index.html')

@app.route('/api/analyze', methods=['POST'])
def analyze():
    """Run analysis on uploaded CSV or sample data"""
    try:
        use_sample = request.form.get('use_sample') == 'true'
        
        if use_sample:
            csv_path = Path(__file__).parent / 'sample_data.csv'
        else:
            if 'file' not in request.files:
                return jsonify({'error': 'No file provided'}), 400
            
            file = request.files['file']
            if not file.filename or file.filename == '':
                return jsonify({'error': 'No file selected'}), 400
            
            if not file.filename.endswith('.csv'):
                return jsonify({'error': 'Only CSV files allowed'}), 400
            
            csv_path = UPLOAD_FOLDER / file.filename
            file.save(csv_path)
        
        # Run analysis
        bot = ChampionCleanersBot(str(csv_path), use_emojis=False)
        results = bot.run_analysis(verbose=False)
        
        if not results:
            return jsonify({'error': 'Analysis failed'}), 500
        
        # Format results for JSON response
        response = {
            'status': 'success',
            'data_summary': results['data_summary'],
            'campaign_metrics': _serialize_metrics(results['campaign_metrics']),
            'detected_issues': results['detected_issues'],
            'recommendations': results['recommendations'],
            'platform_analysis': _serialize_metrics(results['platform_analysis']),
            'device_analysis': _serialize_metrics(results['device_analysis']),
            'budget_allocation': _serialize_metrics(results['budget_allocation'])
        }
        
        return jsonify(response)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def _serialize_metrics(obj):
    """Convert numpy types to Python native types for JSON serialization"""
    if isinstance(obj, dict):
        return {key: _serialize_metrics(val) for key, val in obj.items()}
    elif isinstance(obj, list):
        return [_serialize_metrics(item) for item in obj]
    elif isinstance(obj, (int, float)):
        return float(obj) if isinstance(obj, float) else int(obj)
    elif obj is None or isinstance(obj, (str, bool)):
        return obj
    else:
        return str(obj)

def _serialize_list(items):
    """Serialize a list of items, handling numpy types"""
    if not isinstance(items, list):
        return []
    return [_serialize_metrics(item) for item in items]

@app.route('/api/export-json', methods=['POST'])
def export_json():
    """Export recommendations as JSON file"""
    try:
        data = request.json
        
        output_path = Path(__file__).parent / 'recommendations_export.json'
        with open(output_path, 'w') as f:
            json.dump(data, f, indent=2)
        
        return send_file(output_path, as_attachment=True, download_name='recommendations.json')
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-excel', methods=['POST'])
def export_excel():
    """Export recommendations as Excel file using openpyxl"""
    try:
        if not HAS_OPENPYXL:
            return jsonify({'error': 'Excel export not available'}), 500
        
        data = request.json
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet and create summary sheet
        wb.remove(wb.active)
        
        # Define styles
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        title_font = Font(bold=True, size=14)
        
        # 1. Summary Sheet
        ws = wb.create_sheet('Summary', 0)
        ws['A1'] = 'Keyword Analysis Report'
        ws['A1'].font = title_font
        ws.merge_cells('A1:B1')
        
        row = 3
        summary = data.get('summary', {})
        metrics = [
            ('Total Keywords', summary.get('total_keywords', 0)),
            ('Keywords with Issues', summary.get('keywords_with_issues', 0)),
            ('Lost Opportunities', summary.get('lost_search_opportunities', 0)),
            ('Match Type Conversions', summary.get('match_type_conversions', 0)),
            ('New Keywords Suggested', summary.get('new_keywords_suggested', 0)),
            ('Total Recommendations', summary.get('total_recommendations', 0))
        ]
        
        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws[f'B{row}'].alignment = Alignment(horizontal='center')
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        
        # 2. Keyword Audit Sheet
        ws = wb.create_sheet('Keyword Audit')
        headers = ['Keyword', 'Campaign', 'Issue Type', 'Severity', 'Description']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        audit_items = data.get('keyword_audit', [])
        for row_idx, item in enumerate(audit_items, 2):
            ws.cell(row=row_idx, column=1).value = item.get('keyword', '')
            ws.cell(row=row_idx, column=2).value = item.get('campaign', '')
            ws.cell(row=row_idx, column=3).value = item.get('issue_type', '')
            ws.cell(row=row_idx, column=4).value = item.get('severity', '')
            ws.cell(row=row_idx, column=5).value = item.get('description', '')
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 35
        
        # 3. Lost Searches Sheet
        ws = wb.create_sheet('Lost Searches')
        headers = ['Keyword', 'Campaign', 'Match Type', 'Loss Type', 'Lost Customers', 'Recommendation']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        lost_items = data.get('lost_searches', [])
        for row_idx, item in enumerate(lost_items, 2):
            ws.cell(row=row_idx, column=1).value = item.get('keyword', '')
            ws.cell(row=row_idx, column=2).value = item.get('campaign', '')
            ws.cell(row=row_idx, column=3).value = item.get('match_type', '')
            ws.cell(row=row_idx, column=4).value = item.get('loss_type', '')
            ws.cell(row=row_idx, column=5).value = item.get('potential_searches_lost', 0)
            ws.cell(row=row_idx, column=6).value = item.get('recommendation', '')
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 35
        
        # 4. Match Type Recommendations
        ws = wb.create_sheet('Match Types')
        headers = ['Keyword', 'Campaign', 'Current Type', 'Recommended Type', 'Reason', 'Expected Impact', 'Confidence']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        match_items = data.get('match_recommendations', [])
        for row_idx, item in enumerate(match_items, 2):
            ws.cell(row=row_idx, column=1).value = item.get('keyword', '')
            ws.cell(row=row_idx, column=2).value = item.get('campaign', '')
            ws.cell(row=row_idx, column=3).value = item.get('current_match_type', '')
            ws.cell(row=row_idx, column=4).value = item.get('recommended_match_type', '')
            ws.cell(row=row_idx, column=5).value = item.get('reason', '')
            ws.cell(row=row_idx, column=6).value = item.get('expected_impact', '')
            ws.cell(row=row_idx, column=7).value = item.get('confidence', '')
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 12
        
        # 5. Alignment Sheet
        ws = wb.create_sheet('Alignment')
        headers = ['Keyword', 'Campaign', 'Service', 'Strength', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        alignment_items = data.get('alignment_analysis', [])
        for row_idx, item in enumerate(alignment_items, 2):
            ws.cell(row=row_idx, column=1).value = item.get('keyword', '')
            ws.cell(row=row_idx, column=2).value = item.get('campaign', '')
            ws.cell(row=row_idx, column=3).value = item.get('aligned_service', '')
            strength = item.get('alignment_strength', 0)
            ws.cell(row=row_idx, column=4).value = f"{strength * 100:.0f}%" if strength else "0%"
            ws.cell(row=row_idx, column=5).value = item.get('status', '')
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        
        # 6. Recommendations Sheet
        ws = wb.create_sheet('Recommendations')
        headers = ['Keyword', 'Priority', 'Problem', 'Action', 'Impact']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        recommendations = data.get('top_recommendations', [])
        for row_idx, item in enumerate(recommendations, 2):
            ws.cell(row=row_idx, column=1).value = item.get('keyword', '')
            priority = item.get('priority', '')
            ws.cell(row=row_idx, column=2).value = priority
            ws.cell(row=row_idx, column=3).value = item.get('problem', '')
            ws.cell(row=row_idx, column=4).value = item.get('action', '')
            ws.cell(row=row_idx, column=5).value = item.get('expected_impact', '')
            
            # Color code priority
            if priority == 'High':
                ws.cell(row=row_idx, column=2).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                ws.cell(row=row_idx, column=2).font = Font(color='FFFFFF', bold=True)
            elif priority == 'Medium':
                ws.cell(row=row_idx, column=2).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
                ws.cell(row=row_idx, column=2).font = Font(bold=True)
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 30
        
        # Save to temporary file
        temp_file = Path(__file__).parent / 'uploads' / 'keyword_analysis_report.xlsx'
        wb.save(str(temp_file))
        
        return send_file(
            str(temp_file),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='keyword_analysis_report.xlsx'
        )
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f'Error in export_excel: {str(e)}')
        return jsonify({'error': str(e)}), 500

@app.route('/api/download-sample', methods=['GET'])
def download_sample():
    """Download sample CSV"""
    try:
        sample_path = Path(__file__).parent / 'sample_data.csv'
        return send_file(sample_path, as_attachment=True, download_name='sample_data.csv')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/download-sample-keywords', methods=['GET'])
def download_sample_keywords():
    """Download sample keyword CSV"""
    try:
        sample_path = Path(__file__).parent / 'sample_keywords.csv'
        return send_file(sample_path, as_attachment=True, download_name='sample_keywords.csv')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/analyze-keywords', methods=['POST'])
def analyze_keywords():
    """Run keyword intelligence analysis"""
    try:
        use_sample = request.form.get('use_sample') == 'true'
        
        if use_sample:
            csv_path = Path(__file__).parent / 'sample_keywords.csv'
        else:
            if 'file' not in request.files:
                return jsonify({'error': 'No file provided'}), 400
            
            file = request.files['file']
            if not file.filename or file.filename == '':
                return jsonify({'error': 'No file selected'}), 400
            
            if not file.filename.endswith('.csv'):
                return jsonify({'error': 'Only CSV files allowed'}), 400
            
            csv_path = UPLOAD_FOLDER / file.filename
            file.save(csv_path)
        
        # Run keyword analysis
        if KeywordIntelligenceEngine is None:
            return jsonify({'error': 'Keyword engine not available'}), 500
            
        engine = KeywordIntelligenceEngine()
        
        if not engine.load_keywords(str(csv_path)):
            return jsonify({'error': 'Failed to load keywords'}), 400
        
        if not engine.run_full_analysis():
            return jsonify({'error': 'Analysis failed'}), 500
        
        # Get results summary
        results = engine.get_results_summary()
        
        # Safely extract all fields with defaults
        summary = results.get('summary', {})
        
        # Build match_recommendations
        match_recs = results.get('match_recommendations', [])
        match_recs_serialized = _serialize_list(match_recs[:10] if match_recs else [])
        
        # Build alignment_analysis
        alignment = results.get('alignment_analysis', [])
        alignment_serialized = _serialize_list(alignment)
        
        response = {
            'status': 'success',
            'summary': {
                'total_keywords': int(summary.get('total_keywords', 0)),
                'keywords_with_issues': int(summary.get('keywords_with_issues', 0)),
                'lost_search_opportunities': int(summary.get('lost_search_opportunities', 0)),
                'match_type_conversions': int(summary.get('match_type_conversions_recommended', 0)),
                'new_keywords_suggested': int(summary.get('new_keywords_suggested', 0)),
                'total_recommendations': int(summary.get('total_recommendations', 0))
            },
            'keyword_audit': _serialize_list(results.get('audit', [])[:10]),
            'lost_searches': _serialize_list(results.get('lost_searches', [])[:10]),
            'match_recommendations': match_recs_serialized,
            'alignment_analysis': alignment_serialized,
            'new_keywords': _serialize_list(results.get('new_keywords', [])),
            'service_gaps': _serialize_list(results.get('service_gaps', [])),
            'top_recommendations': _serialize_list(results.get('recommendations', [])[:10])
        }
        
        return jsonify(response)
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/health', methods=['GET'])
def health():
    """Health check endpoint"""
    return jsonify({'status': 'ok', 'version': '1.0.0'})

if __name__ == '__main__':
    print("\n" + "="*80)
    print("CHAMPION CLEANERS BOT - WEB SERVER")
    print("="*80)
    print("\nWeb Server Starting...")
    print("Open your browser and go to:")
    print("   http://localhost:5000")
    print("\nServer running on port 5000")
    print("="*80 + "\n")
    
    # Use Waitress for Windows stability
    try:
        from waitress import serve  # type: ignore
        print("Using Waitress WSGI server (Windows-optimized)")
        print("Server is running - Press CTRL+C to stop\n")
        serve(app, host='127.0.0.1', port=5000, threads=4)
    except ImportError:
        print("Using Flask development server")
        print("Server is running - Press CTRL+C to stop\n")
        app.run(debug=False, host='127.0.0.1', port=5000, threaded=True, use_reloader=False)
    except KeyboardInterrupt:
        print("\n\nServer stopped by user")
    except Exception as e:
        print(f"\nError starting server: {e}")
